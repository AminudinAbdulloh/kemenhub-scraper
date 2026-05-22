"""
=============================================================
 Kemenhub HubUD - Scraper Lalu Lintas Angkutan Udara
=============================================================
 Scrapes air traffic data from:
   https://hubud.kemenhub.go.id/hubud/website/lalu-lintas

 Fitur:
   - Range bulan yang dapat dikonfigurasi
   - Pilihan bandara & kategori (domestik/internasional)
   - Export ke CSV & Excel
   - Logging & resume jika terjadi error
   - Rate limiting agar tidak membebani server
=============================================================
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import random
import logging
import os
import re
import json
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from typing import Optional
import argparse


# ─── Konfigurasi Logging ────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("scraper.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)


# ─── Konstanta ──────────────────────────────────────────────────────────────
BASE_URL  = "https://hubud.kemenhub.go.id/hubud/website/lalu-lintas"

# Pool User-Agent untuk rotasi — mengurangi deteksi sebagai bot
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
]

def get_headers() -> dict:
    """Buat header request dengan User-Agent acak agar terlihat seperti browser berbeda."""
    return {
        "User-Agent": random.choice(USER_AGENTS),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "id-ID,id;q=0.9,en-US;q=0.8,en;q=0.7",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Cache-Control": "max-age=0",
        "Upgrade-Insecure-Requests": "1",
        "Referer": "https://hubud.kemenhub.go.id/",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "same-origin",
        "Sec-Fetch-User": "?1",
    }

# Kolom data yang ada di tabel
DATA_FIELDS = ["Pesawat", "Penumpang Transit", "Penumpang", "Kargo", "Bagasi", "Pos"]


# ─── Fungsi Utilitas ─────────────────────────────────────────────────────────

def clean_number(text: str) -> str:
    """Bersihkan format angka ribuan (1.724 → 1724) dan strip whitespace."""
    if not text:
        return ""
    cleaned = text.strip().replace(".", "").replace(",", "")
    return cleaned if cleaned else "0"


def get_csrf_token(session: requests.Session, retries: int = 5) -> Optional[str]:
    """
    Ambil CSRF token dengan mengakses halaman utama terlebih dahulu.
    Token biasanya ada di hidden input form.
    Otomatis retry dengan cooldown jika terkena 429.
    """
    for attempt in range(1, retries + 1):
        try:
            log.info(f"Mengambil CSRF token dari halaman utama (percobaan {attempt})...")
            resp = session.get(BASE_URL, timeout=30)
            resp.raise_for_status()
            soup = BeautifulSoup(resp.text, "html.parser")

            # Cari input hidden csrf_hubud
            csrf_input = soup.find("input", {"name": "csrf_hubud"})
            if csrf_input:
                token = csrf_input.get("value", "")
                log.info(f"CSRF token ditemukan: {token[:12]}...")
                return token

            # Fallback: cari di meta tag
            meta = soup.find("meta", {"name": "csrf-token"})
            if meta:
                return meta.get("content", "")

            log.warning("CSRF token tidak ditemukan, mencoba tanpa token...")
            return ""

        except requests.exceptions.HTTPError as e:
            status = e.response.status_code if e.response is not None else 0
            if status == 429:
                wait = int(e.response.headers.get("Retry-After", 120))
                wait = max(wait, 120)  # minimal 2 menit
                log.warning(f"429 saat mengambil CSRF token. Menunggu {wait}s (percobaan {attempt}/{retries})...")
                session.cookies.clear()  # hapus cookies lama yang mungkin ter-block
                session.headers.update(get_headers())  # perbarui headers/User-Agent
                time.sleep(wait)
            else:
                log.error(f"Gagal mengambil CSRF token: {e}")
                return None
        except Exception as e:
            log.error(f"Gagal mengambil CSRF token: {e}")
            if attempt < retries:
                time.sleep(10)
            else:
                return None

    log.error("Gagal mendapatkan CSRF token setelah semua percobaan.")
    return None



def scrape_month(
    session: requests.Session,
    csrf_token: str,
    bandara: str,
    period: str,       # format: "YYYY-MM"
    category: str,
    retries: int = 5,
    delay: float = 1.5,
) -> Optional[dict]:
    """
    Ambil data untuk satu bulan tertentu.

    Returns:
        dict berisi data atau None jika gagal.
    """
    params = {
        "csrf_hubud": csrf_token,
        "bandara": bandara,
        "period": period,
        "category": category,
    }

    for attempt in range(1, retries + 1):
        try:
            log.info(f"  ↳ Scraping {period} | bandara={bandara} | kategori={category} (percobaan {attempt})")
            resp = session.get(
                BASE_URL,
                params=params,
                timeout=30,
            )
            resp.raise_for_status()

            soup = BeautifulSoup(resp.text, "html.parser")

            # Cek apakah CSRF token berubah (refresh jika perlu)
            new_csrf = soup.find("input", {"name": "csrf_hubud"})
            if new_csrf:
                csrf_token = new_csrf.get("value", csrf_token)

            # Ambil tabel data
            table = soup.find("table", {"id": "_datatableSimple"})
            if not table:
                log.warning(f"  ✗ Tabel tidak ditemukan untuk periode {period}")
                return None, csrf_token

            # Parse baris data
            tbody = table.find("tbody")
            if not tbody:
                log.warning(f"  ✗ tbody tidak ditemukan untuk periode {period}")
                return None, csrf_token

            rows = tbody.find_all("tr")
            if not rows:
                log.warning(f"  ✗ Tidak ada baris data untuk periode {period}")
                return None, csrf_token

            # Baris pertama adalah header periode (contoh: "May 2026")
            record = {
                "Periode": period,
                "Bandara": bandara.upper(),
                "Kategori": category.capitalize(),
            }

            # Parse setiap baris data
            for row in rows:
                cells = row.find_all("td")
                if len(cells) == 3:
                    label = cells[0].get_text(strip=True).replace("\xa0", " ")
                    datang    = clean_number(cells[1].get_text(strip=True))
                    berangkat = clean_number(cells[2].get_text(strip=True))

                    # Cocokkan label dengan field yang dikenal
                    matched_field = None
                    for field in DATA_FIELDS:
                        if field.lower() in label.lower():
                            matched_field = field
                            break

                    if matched_field:
                        record[f"{matched_field}_Datang"]    = datang
                        record[f"{matched_field}_Berangkat"] = berangkat

            # Pastikan semua field ada (isi 0 jika tidak ada data)
            for field in DATA_FIELDS:
                record.setdefault(f"{field}_Datang",    "0")
                record.setdefault(f"{field}_Berangkat", "0")

            log.info(f"  ✓ Berhasil: {period} — data={len(record)} kolom")
            # Jitter ±30% agar pola request tidak reguler
            jitter = random.uniform(0.7, 1.3)
            time.sleep(delay * jitter)
            return record, csrf_token  # kembalikan juga csrf terbaru

        except requests.exceptions.ConnectionError as e:
            log.warning(f"  ⚠ Koneksi gagal (percobaan {attempt}/{retries}): {e}")
            time.sleep(5 * attempt)
        except requests.exceptions.Timeout:
            log.warning(f"  ⚠ Timeout (percobaan {attempt}/{retries})")
            time.sleep(5 * attempt)
        except requests.exceptions.HTTPError as e:
            status = e.response.status_code if e.response is not None else 0
            if status == 429:
                # Rate limited — cek Retry-After header, atau tunggu default
                retry_after = int(e.response.headers.get("Retry-After", 120))
                retry_after = max(retry_after, 120)  # minimal 2 menit
                log.warning(
                    f"  ⚠ 429 Too Many Requests (percobaan {attempt}/{retries}). "
                    f"Menunggu {retry_after}s sebelum retry..."
                )
                time.sleep(retry_after)
                # Buat fresh session (hapus cookies lama, perbarui User-Agent) dan perbarui CSRF token
                session.cookies.clear()
                session.headers.update(get_headers())
                new_token = get_csrf_token(session)
                if new_token:
                    csrf_token = new_token
                    params["csrf_hubud"] = csrf_token
                # Lanjut ke percobaan berikutnya (jangan break)
            else:
                log.error(f"  ✗ HTTP error: {e}")
                return None, csrf_token
        except Exception as e:
            log.error(f"  ✗ Error tidak terduga: {e}", exc_info=True)
            return None, csrf_token

    return None, csrf_token



def generate_periods(start: str, end: str) -> list[str]:
    """
    Buat daftar periode bulanan antara start dan end (inklusif).

    Args:
        start: "YYYY-MM" atau "YYYY-M"
        end:   "YYYY-MM" atau "YYYY-M"

    Returns:
        list of "YYYY-MM" strings
    """
    try:
        start_dt = datetime.strptime(start.strip(), "%Y-%m")
        end_dt   = datetime.strptime(end.strip(),   "%Y-%m")
    except ValueError as e:
        raise ValueError(f"Format tanggal tidak valid (gunakan YYYY-MM): {e}")

    if start_dt > end_dt:
        raise ValueError("Tanggal mulai tidak boleh lebih besar dari tanggal akhir.")

    periods = []
    current = start_dt
    while current <= end_dt:
        periods.append(current.strftime("%Y-%m"))
        current += relativedelta(months=1)

    return periods


def save_progress(filepath: str, records: list[dict]):
    """Simpan progress ke file JSON sementara."""
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)


def load_progress(filepath: str) -> list[dict]:
    """Muat progress dari file JSON sementara."""
    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


# ─── Fungsi Utama ────────────────────────────────────────────────────────────

def scrape(
    bandara: str  = "UPG",
    category: str = "domestik",
    start: str    = "2020-01",
    end: str      = None,
    output: str   = None,
    delay: float  = 1.5,
    resume: bool  = True,
):
    """
    Scrape data lalu lintas angkutan udara.

    Args:
        bandara  : Kode IATA bandara (contoh: "UPG", "CGK", "SUB")
        category : "domestik" atau "internasional"
        start    : Periode awal  "YYYY-MM"
        end      : Periode akhir "YYYY-MM" (default: bulan saat ini)
        output   : Nama file output (tanpa ekstensi); default auto-generated
        delay    : Jeda antar request (detik)
        resume   : Lanjutkan dari progress sebelumnya jika ada
    """
    # Default end = bulan saat ini
    if end is None:
        end = date.today().strftime("%Y-%m")

    # Nama file output
    if output is None:
        safe_bdr = bandara.upper()
        safe_cat = category.lower()
        output = f"lalinud_{safe_bdr}_{safe_cat}_{start}_{end}"

    progress_file = f"{output}_progress.json"
    csv_file      = f"{output}.csv"
    excel_file    = f"{output}.xlsx"

    log.info("=" * 60)
    log.info("  KEMENHUB HUBUD — SCRAPER LALU LINTAS ANGKUTAN UDARA")
    log.info("=" * 60)
    log.info(f"  Bandara  : {bandara.upper()}")
    log.info(f"  Kategori : {category}")
    log.info(f"  Periode  : {start}  →  {end}")
    log.info(f"  Output   : {csv_file} / {excel_file}")
    log.info("=" * 60)

    # Generate semua periode
    periods = generate_periods(start, end)
    log.info(f"Total periode yang akan diambil: {len(periods)} bulan\n")

    # Muat progress sebelumnya jika ada
    records = []
    scraped_periods = set()
    if resume and os.path.exists(progress_file):
        records = load_progress(progress_file)
        scraped_periods = {r["Periode"] for r in records}
        log.info(f"Resume: {len(records)} periode sudah di-scrape sebelumnya.\n")

    # Buat session HTTP dan set headers awal (termasuk User-Agent konsisten)
    session = requests.Session()
    session.headers.update(get_headers())

    # Ambil CSRF token
    csrf_token = get_csrf_token(session)
    if csrf_token is None:
        log.error("Gagal mendapatkan CSRF token. Program dihentikan.")
        return

    # Iterasi setiap periode
    failed_periods = []
    for i, period in enumerate(periods, 1):
        if period in scraped_periods:
            log.info(f"[{i}/{len(periods)}] Skip {period} (sudah di-scrape)")
            continue

        log.info(f"[{i}/{len(periods)}] Memproses periode {period}...")
        result = scrape_month(
            session=session,
            csrf_token=csrf_token,
            bandara=bandara,
            period=period,
            category=category,
            delay=delay,
        )

        if isinstance(result, tuple):
            data, csrf_token = result
        else:
            data = result

        if data:
            records.append(data)
            scraped_periods.add(period)
            # Simpan progress setiap 5 record
            if len(records) % 5 == 0:
                save_progress(progress_file, records)
        else:
            failed_periods.append(period)
            log.warning(f"  ✗ Gagal mendapatkan data untuk {period}")

    # Simpan progress final
    save_progress(progress_file, records)

    # ─── Export ─────────────────────────────────────────────────────────────
    if not records:
        log.error("Tidak ada data yang berhasil di-scrape.")
        return

    df = pd.DataFrame(records)

    # Urutkan berdasarkan periode
    df = df.sort_values("Periode").reset_index(drop=True)

    # Konversi kolom numerik
    numeric_cols = [
        col for col in df.columns
        if col not in ("Periode", "Bandara", "Kategori")
    ]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col].str.replace(".", "", regex=False), errors="coerce").fillna(0).astype(int)

    # Simpan CSV
    df.to_csv(csv_file, index=False, encoding="utf-8-sig")
    log.info(f"\n[OK] CSV disimpan: {csv_file}")

    # Simpan Excel dengan formatting
    try:
        with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Data")

            ws = writer.sheets["Data"]

            # Format header
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            header_fill   = PatternFill("solid", fgColor="1F4E79")
            header_font   = Font(bold=True, color="FFFFFF", size=11)
            center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border   = Border(
                left   = Side(style="thin"),
                right  = Side(style="thin"),
                top    = Side(style="thin"),
                bottom = Side(style="thin"),
            )

            for cell in ws[1]:
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = center_align
                cell.border    = thin_border

            # Auto-lebar kolom
            for col_idx, col in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_idx)
                max_len    = max(len(str(col)), df[col].astype(str).map(len).max())
                ws.column_dimensions[col_letter].width = min(max_len + 4, 25)

            # Zebra striping
            alt_fill = PatternFill("solid", fgColor="EBF3FB")
            for row_idx in range(2, len(df) + 2):
                for cell in ws[row_idx]:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if row_idx % 2 == 0:
                        cell.fill = alt_fill

        log.info(f"[OK] Excel disimpan: {excel_file}")
    except ImportError:
        log.warning("openpyxl tidak terinstall. Melewati export Excel.")

    # Hapus file progress jika semua berhasil
    if not failed_periods and os.path.exists(progress_file):
        os.remove(progress_file)
        log.info("File progress sementara dihapus.")

    # ─── Ringkasan ──────────────────────────────────────────────────────────
    log.info("\n" + "=" * 60)
    log.info("  RINGKASAN")
    log.info("=" * 60)
    log.info(f"  Total periode diminta : {len(periods)}")
    log.info(f"  Berhasil di-scrape    : {len(records)}")
    log.info(f"  Gagal                 : {len(failed_periods)}")
    if failed_periods:
        log.info(f"  Periode gagal         : {', '.join(failed_periods)}")
    log.info("=" * 60)

    print("\nPreview Data:")
    print(df.to_string(index=False, max_rows=10))

    return df


# ─── CLI Interface ────────────────────────────────────────────────────────────

def parse_args():
    today = date.today().strftime("%Y-%m")

    parser = argparse.ArgumentParser(
        description="Scraper Lalu Lintas Angkutan Udara — Kemenhub HubUD",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Contoh Penggunaan:
  # Scrape UPG domestik dari Jan 2020 sampai sekarang
  python scraper.py --bandara UPG --category domestik --start 2020-01

  # Scrape CGK internasional dari Jan 2023 sampai Des 2024
  python scraper.py --bandara CGK --category internasional --start 2023-01 --end 2024-12

  # Scrape dengan delay lebih lambat (lebih aman)
  python scraper.py --bandara SUB --start 2022-01 --delay 2.0

  # Scrape beberapa bandara sekaligus (jalankan beberapa kali)
  python scraper.py --bandara UPG --start 2020-01
  python scraper.py --bandara CGK --start 2020-01
        """,
    )

    parser.add_argument(
        "--bandara", "-b",
        default="UPG",
        help="Kode IATA bandara (default: UPG). Contoh: CGK, SUB, DPS, KNO",
    )
    parser.add_argument(
        "--category", "-c",
        default="domestik",
        choices=["domestik", "internasional"],
        help="Kategori penerbangan (default: domestik)",
    )
    parser.add_argument(
        "--start", "-s",
        default="2020-01",
        help="Periode awal dalam format YYYY-MM (default: 2020-01)",
    )
    parser.add_argument(
        "--end", "-e",
        default=today,
        help=f"Periode akhir dalam format YYYY-MM (default: bulan ini = {today})",
    )
    parser.add_argument(
        "--output", "-o",
        default=None,
        help="Nama file output tanpa ekstensi (default: auto-generated)",
    )
    parser.add_argument(
        "--delay", "-d",
        type=float,
        default=1.5,
        help="Jeda antar request dalam detik (default: 1.5)",
    )
    parser.add_argument(
        "--no-resume",
        action="store_true",
        help="Jangan gunakan progress yang tersimpan, mulai dari awal",
    )

    return parser.parse_args()


# ─── Entry Point ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    args = parse_args()

    df = scrape(
        bandara  = args.bandara,
        category = args.category,
        start    = args.start,
        end      = args.end,
        output   = args.output,
        delay    = args.delay,
        resume   = not args.no_resume,
    )
    if df is None or len(df) == 0:
        sys.exit(1)
